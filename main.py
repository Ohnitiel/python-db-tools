import argparse
from pathlib import Path

import openpyxl
from openpyxl.utils.cell import get_column_letter

from db_tools.database import DBConnectionRunner

from db_tools.extras import get_available_connections
from db_tools.logger import get_logger, setup_logging

connections = get_available_connections()


def create_arguments() -> argparse.ArgumentParser:
    """
    Creates and configures the argument parser for the command-line interface.

    Returns:
        argparse.ArgumentParser: The configured argument parser.
    """
    parser = argparse.ArgumentParser(
        prog="Database Query Executor",
        description="Conecta em diversas bases de dados e executa queries",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "-c",
        "--connections",
        type=str,
        required=False,
        nargs="+",
        choices=[connections],
        help="Utilizar somente estas conexões. Conexões disponíveis na configuração 'connections'.",
    )
    parser.add_argument("-q", "--query", type=str, required=True)
    parser.add_argument(
        "-s",
        "--save-path",
        type=str,
    )
    parser.add_argument(
        "--environment",
        type=str,
        default="staging",
        choices=["staging", "production", "replica"],
    )
    parser.add_argument("--commit", type=argparse.BooleanOptionalAction, default=False)
    parser.add_argument(
        "--no-parallel", type=argparse.BooleanOptionalAction, default=False
    )
    parser.add_argument("--output-format", type=str, choices=["xlsx", "json", "csv"])
    parser.add_argument(
        "--format-excel", type=argparse.BooleanOptionalAction, default=False
    )
    parser.add_argument(
        "--separate-sheets", type=argparse.BooleanOptionalAction, default=False
    )
    parser.add_argument(
        "--separate-files", type=argparse.BooleanOptionalAction, default=False
    )
    parser.add_argument("--column-name", type=str)
    parser.add_argument("-j", "--max-workers", type=int)
    parser.add_argument(
        "--ignore-cache", type=argparse.BooleanOptionalAction, default=False
    )
    parser.add_argument(
        "--no-cache", type=argparse.BooleanOptionalAction, default=False
    )

    return parser


def validate_args(args: argparse.Namespace):
    """
    Validates the command-line arguments.

    Args:
        args: The parsed command-line arguments.
    """
    if args.separate_sheets and args.separate_files:
        # TODO Add a picker/raise error/prioritize one?
        pass

    if args.save_path is not None:
        args.save_path = Path(args.save_path)
        if not args.save_path.parent.exists():
            print("Diretório informado não existe!")
            response = input("Deseja criá-lo?")
            if response.lower() == "s":
                args.save_path.parent.mkdirs(exist_ok=True, parents=True)


def format_excel_file(file_path: Path, datetime_columns: list[int]):
    """Apply formatting to Excel file."""
    try:
        wb = openpyxl.load_workbook(file_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            for col in datetime_columns:
                column_letter = get_column_letter(col)
                for cell in ws[column_letter]:
                    cell.number_format = "dd/mm/yyyy"
        wb.save(file_path)
    except Exception as e:
        logger.error(f"Failed to format Excel file: {e}")


def main():
    """
    The main function of the application.
    """
    parser = create_arguments()
    args = parser.parse_args()
    validate_args(args)
    runner = DBConnectionRunner(
        args.environment,
        args.connections,
        args.max_workers,
        args.save_path,
        args.output_format,
    )
    add_connection_column = True if args.column_name else False
    try:
        df = runner.execute_query_multi_db(
            args.query,
            args.commit,
            args.no_parallel,
            add_connection_column,
            args.column_name,
            args.no_cache,
            args.ignore_cache,
        )
        if not df.empty and args.format_excel_output:
            dtypes_df = df.dtypes.reset_index(drop=True).astype(str)
            datetimecols = dtypes_df[dtypes_df.str.contains("datetime")].index

            format_excel_file(args.save_path, datetimecols)
    finally:
        runner.close_all()


if __name__ == "__main__":
    from dotenv import load_dotenv

    load_dotenv()

    setup_logging()
    logger = get_logger("db_tools")

    main()
