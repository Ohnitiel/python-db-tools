import re
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# TODO Handle Excel max row and column size
XL_MAX_ROWS: int = 1_048_575
XL_MAX_COLS: int = 16_384


def format_excel(wb: Workbook, df: pd.DataFrame):
    datetime_columns = [
        i
        for i, dtype in enumerate(df.dtypes)
        if pd.api.types.is_datetime64_any_dtype(dtype)
    ]

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        if ws.max_row <= 1:
            continue

        last_column = get_column_letter(ws.max_column)
        last_cell = f"{last_column}{ws.max_row}"

        table_name = re.sub(r"[^\w]", "_", sheet)
        table = Table(displayName=table_name, ref=f"A1:{last_cell}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(table)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        for idx in datetime_columns:
            column_letter = get_column_letter(idx + 1)
            for cell in ws[column_letter]:
                if cell.value:
                    cell.number_format = "dd/mm/yyyy"


def export_data(
    save_path: Path,
    df: pd.DataFrame,
    file_format: str,
    single_file: bool,
    single_sheet: bool,
    connection_column: Optional[str] = None,
    format: bool = True,
):
    if (not single_file or not single_sheet) and connection_column is None:
        raise ValueError(
            "connection_column is required when single_file or single_sheet are False!"
        )

    if connection_column and connection_column not in df.columns:
        raise ValueError(f"{connection_column} not found in Dataframe!")

    df = df.infer_objects()

    datetime_columns = [
        i
        for i, dtype in enumerate(df.dtypes)
        if pd.api.types.is_datetime64_any_dtype(dtype)
    ]

    for i in datetime_columns:
        df[df.columns[i]] = df[df.columns[i]].dt.tz_convert(None)

    if file_format == "xlsx":
        if single_file and single_sheet:
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Data", index=False)
                if format:
                    format_excel(writer.book, df)

        elif single_file:
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                for connection in df[connection_column].unique():
                    conn_df = df[df[connection_column] == connection]
                    conn_df = conn_df.drop(columns=[connection_column])
                    conn_df.to_excel(writer, sheet_name=connection, index=False)
                if format:
                    format_excel(writer.book, df)

        elif single_sheet:
            for connection in df[connection_column].unique():
                file_path = save_path.with_stem(f"{save_path.stem}_{connection}")
                with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                    conn_df = df[df[connection_column] == connection]
                    conn_df = conn_df.drop(columns=[connection_column])
                    conn_df.to_excel(writer, index=False)
                    if format:
                        format_excel(writer.book, df)
    elif file_format == "json":
        df.to_json(save_path, orient="records", indent=4)
    elif file_format == "csv":
        df.to_csv(save_path, index=False)